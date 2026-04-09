import re
from io import BytesIO

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Conciliador NIP vs Agente", layout="wide")


# =========================
# Helpers
# =========================
def normalizar_texto(x):
    if pd.isna(x):
        return ""
    return str(x).strip().upper()


def normalizar_importe(x):
    if pd.isna(x):
        return 0.0

    if isinstance(x, (int, float)):
        return float(x)

    s = str(x).strip()

    if s == "":
        return 0.0

    s = s.replace("USD", "").replace("US$", "").replace("$", "").replace(" ", "")

    if "(" in s and ")" in s:
        s = "-" + s.replace("(", "").replace(")", "")

    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s and "." not in s:
        partes = s.split(",")
        if len(partes[-1]) in (1, 2):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")

    try:
        return float(s)
    except:
        return 0.0


def extraer_sis(texto):
    texto = normalizar_texto(texto)
    encontrados = re.findall(r"SIS\d{8,}", texto)
    return encontrados


def es_sis_multiple(texto):
    return len(extraer_sis(texto)) > 1


def primer_sis(texto):
    encontrados = extraer_sis(texto)
    return encontrados[0] if encontrados else ""


def buscar_columna(df, aliases, obligatoria=True):
    cols_norm = {normalizar_texto(c): c for c in df.columns}
    for alias in aliases:
        alias_norm = normalizar_texto(alias)
        if alias_norm in cols_norm:
            return cols_norm[alias_norm]

    if obligatoria:
        raise ValueError(f"No encontré ninguna de estas columnas: {aliases}")
    return None


def detectar_columna_sis_directa(df):
    patrones = [
        "NRO OPERACION", "OPERACION", "NRO OPERACIÓN", "SIS", "NRO SIS",
        "N° SIS", "NRO DE OPERACION", "NRO DE OPERACIÓN", "OPERACIÓN"
    ]
    for p in patrones:
        col = buscar_columna(df, [p], obligatoria=False)
        if col:
            return col
    return None


# =========================
# Preparación de hojas
# =========================
def preparar_hoja_agente(df):
    col_job = buscar_columna(
        df,
        ["JOB NO", "JOB", "COMPROBANTE", "FACTURA", "NRO JOB", "JOB NUMBER"],
        obligatoria=False
    )
    col_mbl = buscar_columna(
        df,
        ["MBL", "MASTER", "MASTER BL", "MASTER BILL"],
        obligatoria=False
    )
    col_sis = buscar_columna(
        df,
        ["NRO SIS", "SIS", "N° SIS", "NUMERO SIS", "NRO OPERACION", "OPERACION"]
    )
    col_total = buscar_columna(
        df,
        ["TOTAL", "IMPORTE", "TOTAL AGENTE", "SALDO", "MONTO"]
    )

    out = pd.DataFrame()
    out["JOB NO"] = df[col_job] if col_job else ""
    out["MBL"] = df[col_mbl] if col_mbl else ""
    out["NRO SIS ORIGINAL"] = df[col_sis]
    out["NRO SIS (Agente)"] = df[col_sis].apply(primer_sis)
    out["TOTAL Agente"] = df[col_total].apply(normalizar_importe)

    out["JOB NO"] = out["JOB NO"].apply(normalizar_texto)
    out["MBL"] = out["MBL"].apply(normalizar_texto)
    out["NRO SIS ORIGINAL"] = out["NRO SIS ORIGINAL"].apply(normalizar_texto)
    out["NRO SIS (Agente)"] = out["NRO SIS (Agente)"].apply(normalizar_texto)

    return out


def preparar_hoja_nip(df):
    def buscar_sis_en_fila(row):
        for val in row:
            if pd.isna(val):
                continue
            encontrados = re.findall(r"SIS\d{8,}", str(val).upper())
            if encontrados:
                return encontrados[0]
        return ""

    col_total = buscar_columna(df, ["TOTAL", "IMPORTE", "TOTAL NIP", "SALDO", "MONTO"])

    col_op = detectar_columna_sis_directa(df)
    col_comp = buscar_columna(df, ["COMPROBANTE", "FACTURA", "NRO FACTURA"], obligatoria=False)
    col_master = buscar_columna(df, ["MASTER", "MBL", "MASTER BL"], obligatoria=False)
    col_house = buscar_columna(df, ["HOUSE", "HBL"], obligatoria=False)
    col_fecha = buscar_columna(df, ["FECHA EMISION", "FECHA", "EMISION"], obligatoria=False)
    col_moneda = buscar_columna(df, ["MONEDA"], obligatoria=False)

    out = pd.DataFrame()

    if col_op:
        out["NRO OPERACION"] = df[col_op].apply(lambda x: primer_sis(x) if primer_sis(x) else normalizar_texto(x))
    else:
        out["NRO OPERACION"] = df.apply(buscar_sis_en_fila, axis=1)

    out["COMPROBANTE"] = df[col_comp].apply(normalizar_texto) if col_comp else ""
    out["MASTER"] = df[col_master].apply(normalizar_texto) if col_master else ""
    out["HOUSE"] = df[col_house].apply(normalizar_texto) if col_house else ""
    out["FECHA EMISION"] = df[col_fecha] if col_fecha else ""
    out["MONEDA"] = df[col_moneda].apply(normalizar_texto) if col_moneda else ""
    out["TOTAL NIP"] = df[col_total].apply(normalizar_importe)

    out["NRO OPERACION"] = out["NRO OPERACION"].apply(normalizar_texto)

    return out


# =========================
# Conciliación
# =========================
def conciliar(df_agente_raw, df_nip_raw):
    agente = preparar_hoja_agente(df_agente_raw)
    nip = preparar_hoja_nip(df_nip_raw)

    revisar_manual = agente[agente["NRO SIS ORIGINAL"].apply(es_sis_multiple)].copy()
    revisar_manual["NOTA"] = "Buscar manualmente en Hoja 2 por BL o sufijos del SIS"

    agente_simple = agente[~agente["NRO SIS ORIGINAL"].apply(es_sis_multiple)].copy()
    agente_simple = agente_simple[agente_simple["NRO SIS (Agente)"] != ""].copy()

    nip_valid = nip[nip["NRO OPERACION"] != ""].copy()

    nip_group = (
        nip_valid.groupby("NRO OPERACION", as_index=False)
        .agg({
            "COMPROBANTE": lambda x: " | ".join(sorted({str(v) for v in x if str(v).strip() != ""})),
            "MASTER": lambda x: " | ".join(sorted({str(v) for v in x if str(v).strip() != ""})),
            "HOUSE": lambda x: " | ".join(sorted({str(v) for v in x if str(v).strip() != ""})),
            "FECHA EMISION": "first",
            "MONEDA": "first",
            "TOTAL NIP": "sum"
        })
    )

    base = agente_simple.merge(
        nip_group,
        left_on="NRO SIS (Agente)",
        right_on="NRO OPERACION",
        how="left"
    )

    base["TOTAL NIP"] = base["TOTAL NIP"].fillna(0.0)
    base["DIFERENCIA"] = (base["TOTAL Agente"].abs() - base["TOTAL NIP"].abs()).round(2)

    coincidentes_ok = base[
        (base["NRO OPERACION"].notna()) &
        (base["NRO OPERACION"] != "") &
        (base["DIFERENCIA"].abs() <= 0.01)
    ].copy()

    con_diferencia = base[
        (base["NRO OPERACION"].notna()) &
        (base["NRO OPERACION"] != "") &
        (base["DIFERENCIA"].abs() > 0.01)
    ].copy()

    solo_agente = base[
        (base["NRO OPERACION"].isna()) | (base["NRO OPERACION"] == "")
    ].copy()

    usados = set(base["NRO OPERACION"].dropna().astype(str))
    solo_nip = nip_valid[~nip_valid["NRO OPERACION"].astype(str).isin(usados)].copy()

    coincidentes_ok_final = coincidentes_ok[
        ["JOB NO", "MBL", "NRO SIS (Agente)", "NRO OPERACION", "TOTAL Agente", "TOTAL NIP", "DIFERENCIA"]
    ].copy()
    coincidentes_ok_final["NOTA"] = ""

    con_diferencia_final = con_diferencia[
        ["JOB NO", "MBL", "NRO SIS (Agente)", "NRO OPERACION", "TOTAL Agente", "TOTAL NIP", "DIFERENCIA", "COMPROBANTE"]
    ].copy()
    con_diferencia_final.rename(columns={"COMPROBANTE": "COMPROBANTES NIP"}, inplace=True)
    con_diferencia_final["NOTA"] = ""

    solo_agente_final = solo_agente[
        ["JOB NO", "MBL", "NRO SIS (Agente)", "TOTAL Agente"]
    ].copy()
    solo_agente_final["NOTA"] = ""

    solo_nip_final = solo_nip[
        ["NRO OPERACION", "COMPROBANTE", "MASTER", "HOUSE", "FECHA EMISION", "MONEDA", "TOTAL NIP"]
    ].copy()

    revisar_manual_final = revisar_manual[
        ["JOB NO", "MBL", "NRO SIS ORIGINAL", "TOTAL Agente", "NOTA"]
    ].copy()
    revisar_manual_final.rename(columns={"NRO SIS ORIGINAL": "NRO SIS COMPLETO (Agente)"}, inplace=True)

    resumen = {
        "reg_agente": int(len(agente)),
        "imp_agente": round(float(agente["TOTAL Agente"].sum()), 2),
        "reg_nip": int(len(nip_valid)),
        "imp_nip": round(float(nip_valid["TOTAL NIP"].sum()), 2),
        "ok_reg": int(len(coincidentes_ok_final)),
        "ok_imp_ag": round(float(coincidentes_ok_final["TOTAL Agente"].sum()), 2) if not coincidentes_ok_final.empty else 0.0,
        "ok_imp_nip": round(float(coincidentes_ok_final["TOTAL NIP"].sum()), 2) if not coincidentes_ok_final.empty else 0.0,
        "dif_reg": int(len(con_diferencia_final)),
        "dif_imp_ag": round(float(con_diferencia_final["TOTAL Agente"].sum()), 2) if not con_diferencia_final.empty else 0.0,
        "dif_imp_nip": round(float(con_diferencia_final["TOTAL NIP"].sum()), 2) if not con_diferencia_final.empty else 0.0,
        "solo_ag_reg": int(len(solo_agente_final)),
        "solo_ag_imp": round(float(solo_agente_final["TOTAL Agente"].sum()), 2) if not solo_agente_final.empty else 0.0,
        "solo_nip_reg": int(len(solo_nip_final)),
        "solo_nip_imp": round(float(solo_nip_final["TOTAL NIP"].sum()), 2) if not solo_nip_final.empty else 0.0,
        "rev_reg": int(len(revisar_manual_final)),
        "rev_imp": round(float(revisar_manual_final["TOTAL Agente"].sum()), 2) if not revisar_manual_final.empty else 0.0,
    }

    return {
        "resumen": resumen,
        "coincidentes_ok": coincidentes_ok_final,
        "con_diferencia": con_diferencia_final,
        "solo_agente": solo_agente_final,
        "solo_nip": solo_nip_final,
        "revisar_manual": revisar_manual_final,
    }


# =========================
# Excel output
# =========================
def aplicar_estilo_titulo(ws, rango, fill_color="1F4E78"):
    for row in ws[rango]:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")


def auto_ancho(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(val))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)


def escribir_dataframe(ws, df, start_row, start_col=1, include_header=True):
    row_idx = start_row
    col_idx = start_col

    if include_header:
        for j, col_name in enumerate(df.columns, start=col_idx):
            ws.cell(row=row_idx, column=j, value=col_name)
        row_idx += 1

    for _, row in df.iterrows():
        for j, value in enumerate(row.tolist(), start=col_idx):
            ws.cell(row=row_idx, column=j, value=value)
        row_idx += 1

    return row_idx


def generar_excel(resultado):
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"

    resumen = resultado["resumen"]
    ok = resultado["coincidentes_ok"]
    dif = resultado["con_diferencia"]
    solo_ag = resultado["solo_agente"]
    solo_nip = resultado["solo_nip"]
    rev = resultado["revisar_manual"]

    ws["A1"] = "CONCILIACIÓN DE CUENTAS — UNION CARGO INTERNATIONAL"
    ws["A2"] = "Cuenta Agente (Hoja 1) vs Cuenta NIP (Hoja 2)"
    ws["A3"] = "Match por NRO SIS | SIS múltiple → Revisar Manual"

    ws["A5"] = "CONCEPTO"
    ws["B5"] = "REGISTROS"
    ws["C5"] = "IMPORTE (USD)"
    aplicar_estilo_titulo(ws, "A5:C5")

    resumen_rows = [
        ["Cuenta Agente (Hoja 1)", resumen["reg_agente"], resumen["imp_agente"]],
        ["Cuenta NIP (Hoja 2)", resumen["reg_nip"], resumen["imp_nip"]],
        ["Coincidentes OK", resumen["ok_reg"], resumen["ok_imp_ag"]],
        ["Con Diferencia", resumen["dif_reg"], resumen["dif_imp_ag"]],
        ["Solo Agente", resumen["solo_ag_reg"], resumen["solo_ag_imp"]],
        ["Solo NIP", resumen["solo_nip_reg"], resumen["solo_nip_imp"]],
        ["Revisar Manual", resumen["rev_reg"], resumen["rev_imp"]],
    ]

    fila = 6
    for r in resumen_rows:
        ws.cell(fila, 1, r[0])
        ws.cell(fila, 2, r[1])
        ws.cell(fila, 3, r[2])
        fila += 1

    ws["A14"] = "Control"
    ws["B14"] = "Diferencia magnitudes Agente vs NIP"
    ws["C14"] = round(abs(resumen["imp_agente"]) - abs(resumen["imp_nip"]), 2)
    aplicar_estilo_titulo(ws, "A14:C14", fill_color="6B7280")

    auto_ancho(ws)

    ws_ok = wb.create_sheet("Coincidentes OK")
    ws_ok["A1"] = f"COINCIDENTES SIN DIFERENCIA — {len(ok)} registros"
    escribir_dataframe(ws_ok, ok, start_row=2)
    if len(ok.columns) > 0:
        aplicar_estilo_titulo(ws_ok, f"A2:{get_column_letter(len(ok.columns))}2")
    auto_ancho(ws_ok)

    ws_dif = wb.create_sheet("Con Diferencia")
    ws_dif["A1"] = f"COINCIDENTES CON DIFERENCIA DE IMPORTE — {len(dif)} registros"
    escribir_dataframe(ws_dif, dif, start_row=2)
    if len(dif.columns) > 0:
        aplicar_estilo_titulo(ws_dif, f"A2:{get_column_letter(len(dif.columns))}2")
    auto_ancho(ws_dif)

    ws_sa = wb.create_sheet("Solo Agente")
    ws_sa["A1"] = f"SOLO EN CUENTA AGENTE — {len(solo_ag)} registros"
    escribir_dataframe(ws_sa, solo_ag, start_row=2)
    if len(solo_ag.columns) > 0:
        aplicar_estilo_titulo(ws_sa, f"A2:{get_column_letter(len(solo_ag.columns))}2")
    total_row = len(solo_ag) + 3
    ws_sa.cell(total_row, 1, "TOTAL")
    if len(solo_ag) > 0:
        ws_sa.cell(total_row, 4, f"=SUM(D3:D{total_row-1})")
    auto_ancho(ws_sa)

    ws_sn = wb.create_sheet("Solo NIP")
    ws_sn["A1"] = f"SOLO EN CUENTA NIP — {len(solo_nip)} registros"
    escribir_dataframe(ws_sn, solo_nip, start_row=2)
    if len(solo_nip.columns) > 0:
        aplicar_estilo_titulo(ws_sn, f"A2:{get_column_letter(len(solo_nip.columns))}2")
    auto_ancho(ws_sn)

    ws_rm = wb.create_sheet("Revisar Manual")
    ws_rm["A1"] = f"REVISAR MANUALMENTE — {len(rev)} registros con SIS múltiple"
    ws_rm["A2"] = "Tienen múltiples NRO SIS agrupados. Verificar manualmente en Hoja 2 por BL o sufijos."
    escribir_dataframe(ws_rm, rev, start_row=3)
    if len(rev.columns) > 0:
        aplicar_estilo_titulo(ws_rm, f"A3:{get_column_letter(len(rev.columns))}3")
    total_row_rm = len(rev) + 4
    ws_rm.cell(total_row_rm, 1, "TOTAL")
    if len(rev) > 0:
        ws_rm.cell(total_row_rm, 4, f"=SUM(D4:D{total_row_rm-1})")
    auto_ancho(ws_rm)

    for hoja in [ws, ws_ok, ws_dif, ws_sa, ws_sn, ws_rm]:
        for row in hoja.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# =========================
# UI
# =========================
st.title("Conciliador mensual NIP vs Agente")
st.write("Subí un Excel con 2 hojas: hoja 1 = Agente, hoja 2 = NIP.")

archivo = st.file_uploader("Subir archivo base", type=["xlsx"])

with st.expander("Formato esperado"):
    st.markdown("""
**Hoja 1 (Agente)** debe tener alguna columna equivalente a:
- JOB NO
- MBL
- NRO SIS
- TOTAL

**Hoja 2 (NIP)** debe tener al menos:
- TOTAL

Y si no tiene una columna llamada NRO OPERACION, la app intenta encontrar el SIS automáticamente dentro de cualquier columna.
""")

if archivo:
    try:
        xls = pd.ExcelFile(archivo)

        if len(xls.sheet_names) < 2:
            st.error("El archivo debe tener al menos 2 hojas.")
            st.stop()

        hoja_agente = st.selectbox("Seleccioná la hoja del Agente", xls.sheet_names, index=0)
        hoja_nip = st.selectbox("Seleccioná la hoja de NIP", xls.sheet_names, index=1)

        df_agente = pd.read_excel(archivo, sheet_name=hoja_agente)
        df_nip = pd.read_excel(archivo, sheet_name=hoja_nip)

        if st.button("Procesar conciliación", type="primary"):
            resultado = conciliar(df_agente, df_nip)

            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Coincidentes OK", len(resultado["coincidentes_ok"]))
            c2.metric("Con diferencia", len(resultado["con_diferencia"]))
            c3.metric("Solo Agente", len(resultado["solo_agente"]))
            c4.metric("Solo NIP", len(resultado["solo_nip"]))

            st.subheader("Vista previa")

            st.write("Coincidentes OK")
            st.dataframe(resultado["coincidentes_ok"], use_container_width=True)

            st.write("Con Diferencia")
            st.dataframe(resultado["con_diferencia"], use_container_width=True)

            st.write("Solo Agente")
            st.dataframe(resultado["solo_agente"], use_container_width=True)

            st.write("Solo NIP")
            st.dataframe(resultado["solo_nip"], use_container_width=True)

            st.write("Revisar Manual")
            st.dataframe(resultado["revisar_manual"], use_container_width=True)

            excel_final = generar_excel(resultado)

            st.download_button(
                label="Descargar conciliación final",
                data=excel_final.getvalue(),
                file_name="Conciliacion_Union_generada.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    except Exception as e:
        st.error(f"Hubo un problema al procesar el archivo: {e}")
